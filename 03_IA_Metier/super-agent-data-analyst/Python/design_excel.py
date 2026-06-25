import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule

input_csv = 'Fichier_Prospection_Unifie.csv'
output_excel = 'Prospection_Design.xlsx'

print("Lecture du fichier CSV...")
df = pd.read_csv(input_csv)

# Création d'un nouveau classeur Excel
wb = Workbook()
ws = wb.active
ws.title = "Prospection"

# Définition des styles (Thème Prestige / Professionnel)
header_fill = PatternFill(start_color="1A1E26", end_color="1A1E26", fill_type="solid")
header_font = Font(name="Montserrat", color="D4AF37", bold=True, size=11)
align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
thin_border = Border(left=Side(style='thin', color='E0E0E0'), 
                     right=Side(style='thin', color='E0E0E0'), 
                     top=Side(style='thin', color='E0E0E0'), 
                     bottom=Side(style='thin', color='E0E0E0'))

# Ajout des données au classeur
for r in dataframe_to_rows(df, index=False, header=True):
    ws.append(r)

# Application des styles à l'en-tête
for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = align_center

# Application des styles aux cellules de données
for row in ws.iter_rows(min_row=2, max_col=ws.max_column, max_row=ws.max_row):
    for cell in row:
        cell.alignment = align_left
        cell.border = thin_border
        
        # Mettre en évidence les téléphones
        if cell.column == 2: # Colonne Téléphone
            cell.font = Font(bold=True)

# Ajustement de la largeur des colonnes
col_widths = {'A': 40, 'B': 18, 'C': 35, 'D': 25, 'E': 45, 'F': 25, 'G': 60}
for col, width in col_widths.items():
    ws.column_dimensions[col].width = width

# Ajout d'un système de filtre automatique
ws.auto_filter.ref = ws.dimensions

# Règles de mise en forme conditionnelle pour la colonne 'Statut_Appel' (Colonne E)
red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
red_font = Font(color='9C0006')
green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
green_font = Font(color='006100')
grey_fill = PatternFill(start_color='EAEAEA', end_color='EAEAEA', fill_type='solid')
grey_font = Font(color='555555')
black_fill = PatternFill(start_color='000000', end_color='000000', fill_type='solid')
white_font = Font(color='FFFFFF')

# Application de la mise en forme conditionnelle (Colonne Statut est la E -> lettre E)
statut_col_letter = 'E'
statut_range = f"{statut_col_letter}2:{statut_col_letter}{ws.max_row}"

ws.conditional_formatting.add(statut_range, CellIsRule(operator='equal', formula=['"NRP"'], stopIfTrue=True, fill=grey_fill, font=grey_font))
ws.conditional_formatting.add(statut_range, CellIsRule(operator='equal', formula=['"REFUS CATÉGORIQUE"'], stopIfTrue=True, fill=red_fill, font=red_font))
ws.conditional_formatting.add(statut_range, CellIsRule(operator='equal', formula=['"REFUS ARGUMENTÉ"'], stopIfTrue=True, fill=red_fill, font=red_font))
ws.conditional_formatting.add(statut_range, CellIsRule(operator='equal', formula=['"SUIVI NÉGO"'], stopIfTrue=True, fill=green_fill, font=green_font))
ws.conditional_formatting.add(statut_range, CellIsRule(operator='equal', formula=['"OPPOSITION RGPD"'], stopIfTrue=True, fill=black_fill, font=white_font))


wb.save(output_excel)
print(f"Fichier Excel design généré avec succès : {output_excel}")
